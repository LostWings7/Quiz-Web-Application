import csv
import io
import re

DEFAULT_FILLS = ('00000000', 'FFFFFFFF', 'FF000000', '000000', 'FFFFFF', None, 'auto', 'indexed')


def column_label(index, headers=None, has_header=False):
    if has_header and headers and index < len(headers):
        label = str(headers[index]).strip()
        if label:
            return label
    return f'Column {index + 1}'


def cell_has_highlight(cell):
    fill = cell.fill
    if fill is None or fill.fill_type in (None, 'none'):
        return False
    color = fill.start_color
    if color is None:
        return False
    rgb = getattr(color, 'rgb', None)
    if rgb and str(rgb).upper() not in {str(v).upper() for v in DEFAULT_FILLS if v}:
        return True
    indexed = getattr(color, 'indexed', None)
    if indexed not in (None, 64):
        return True
    theme = getattr(color, 'theme', None)
    return theme is not None


import os
import uuid
import zipfile
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage


def parse_uploaded_file(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith('.zip'):
        return parse_zip(uploaded_file)
    if name.endswith('.xlsx'):
        return parse_xlsx(uploaded_file)
    if name.endswith('.xls'):
        raise ValueError('Legacy .xls files are not supported. Please save as .xlsx or .csv.')
    return parse_csv(uploaded_file)


def parse_zip(uploaded_file):
    try:
        z = zipfile.ZipFile(uploaded_file)
    except zipfile.BadZipFile:
        raise ValueError('Invalid or corrupted ZIP archive.')

    names = z.namelist()
    spreadsheet_name = next(
        (n for n in names if not n.startswith('__MACOSX') and (n.lower().endswith('.xlsx') or n.lower().endswith('.csv'))),
        None,
    )
    if not spreadsheet_name:
        raise ValueError('No .xlsx or .csv spreadsheet file found inside the ZIP package.')

    extracted_images = {}
    valid_exts = ('.png', '.jpg', '.jpeg', '.gif', '.webp')
    import_subfolder = f"imported_images/{uuid.uuid4().hex[:8]}"

    for name in names:
        if name.startswith('__MACOSX') or name.endswith('/'):
            continue
        lower_name = name.lower()
        if any(lower_name.endswith(ext) for ext in valid_exts):
            basename = os.path.basename(name)
            if basename:
                content = z.read(name)
                save_path = f"{import_subfolder}/{basename}"
                saved_name = default_storage.save(save_path, ContentFile(content))
                extracted_images[basename] = saved_name
                extracted_images[basename.lower()] = saved_name

    sheet_data = z.read(spreadsheet_name)
    sheet_file = ContentFile(sheet_data, name=os.path.basename(spreadsheet_name))

    if spreadsheet_name.lower().endswith('.xlsx'):
        parsed = parse_xlsx(sheet_file)
    else:
        parsed = parse_csv(sheet_file)

    parsed['extracted_images'] = extracted_images
    parsed['filename'] = uploaded_file.name
    return parsed


def parse_csv(uploaded_file):
    raw = uploaded_file.read()
    uploaded_file.seek(0)
    for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            decoded = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            decoded = None
    if decoded is None:
        raise ValueError('Could not read the CSV file encoding.')

    reader = csv.reader(io.StringIO(decoded))
    rows = [list(row) for row in reader if any(str(cell).strip() for cell in row)]
    if not rows:
        raise ValueError('The file is empty.')

    max_cols = max(len(row) for row in rows)
    normalized = []
    for row in rows:
        padded = list(row) + [''] * (max_cols - len(row))
        normalized.append([str(cell).strip() for cell in padded])

    return {
        'rows': normalized,
        'highlighted': [],
        'supports_color': False,
        'filename': uploaded_file.name,
    }


def parse_xlsx(uploaded_file):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('Excel support requires openpyxl. Contact your administrator.') from exc

    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        raise ValueError('The spreadsheet is empty.')

    rows = []
    highlighted = []
    for r in range(1, max_row + 1):
        row_values = []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            value = cell.value
            row_values.append('' if value is None else str(value).strip())
            if cell_has_highlight(cell):
                highlighted.append([r - 1, c - 1])
        if any(row_values):
            rows.append(row_values)

    if not rows:
        raise ValueError('The spreadsheet is empty.')

    max_cols = max(len(row) for row in rows)
    normalized = [row + [''] * (max_cols - len(row)) for row in rows]

    return {
        'rows': normalized,
        'highlighted': highlighted,
        'supports_color': True,
        'filename': uploaded_file.name,
    }


def build_column_choices(num_columns, headers=None, has_header=False):
    return [
        {'index': i, 'label': column_label(i, headers, has_header)}
        for i in range(num_columns)
    ]


def preview_rows(rows, has_header, limit=5):
    if has_header and rows:
        return rows[1:limit + 1]
    return rows[:limit]


def preview_headers(rows, has_header):
    if has_header and rows:
        return rows[0]
    return None


def cell_value(row, col_index):
    if col_index is None or col_index == '':
        return ''
    try:
        col_index = int(col_index)
    except (TypeError, ValueError):
        return ''
    if col_index < 0 or col_index >= len(row):
        return ''
    return str(row[col_index]).strip()


def highlighted_col_for_row(highlighted, row_index, option_columns):
    highlighted_cols = {int(c) for r, c in highlighted if int(r) == row_index}
    for opt_idx, col_index in enumerate(option_columns):
        if col_index is not None and col_index != '' and int(col_index) in highlighted_cols:
            return opt_idx
    return None


def parse_order_value(raw):
    if raw is None or raw == '':
        return 0
    try:
        return max(0, int(float(str(raw))))
    except (ValueError, TypeError):
        return 0


import urllib.request
import urllib.parse

def download_image_from_url(url):
    """
    Downloads an image from a URL and saves it to media/imported_images/.
    Returns the relative path in storage if successful, else None.
    """
    url_str = str(url).strip()
    if not (url_str.startswith('http://') or url_str.startswith('https://')):
        return None
    try:
        req = urllib.request.Request(
            url_str,
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        )
        with urllib.request.urlopen(req, timeout=5) as response:
            content = response.read()
            parsed = urllib.parse.urlparse(url_str)
            filename = os.path.basename(parsed.path)
            if not filename or '.' not in filename:
                filename = f"downloaded_{uuid.uuid4().hex[:8]}.png"

            save_path = f"imported_images/{uuid.uuid4().hex[:8]}/{filename}"
            saved_name = default_storage.save(save_path, ContentFile(content))
            return saved_name
    except Exception as e:
        print(f"Failed to download image from {url_str}: {e}")
        return None

